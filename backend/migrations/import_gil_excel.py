"""
import_gil_excel.py  (v2 – rewritten for actual local DB schema)
=================================================================
Imports GIL settlement data from banking_settlement_*.xlsx files
into monthly_banking_settlement + tod_daily_summary.

Actual local DB schema (verified 2026-07-22):
  - tenants.id = 2  (GIL)
  - plants.id = 100 (GIL_MH_HYBRID)
  - plant_energy_sources.id = 100 (WIND, used as combined source)
  - tod_slot_definitions: NORMAL=7, OFF_PEAK=8, PEAK=5  (discom_id=2)
  - monthly_banking_settlement uses plant_energy_source_id + consumption_unit_id (0 for GIL)
  - tod_daily_summary.consumption_unit_id is nullable

Usage:
    python import_gil_excel.py [--dry-run] [--dir path/to/gil/files]
"""

import os, sys, argparse, datetime, re
from decimal import Decimal
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── CONFIG ─────────────────────────────────────────────────────────────────────

DB_CONFIG = dict(
    host     = os.getenv("DB_HOST",     "localhost"),
    port     = int(os.getenv("DB_PORT", "5432")),
    user     = os.getenv("DB_USER",     "integrum"),
    password = os.getenv("DB_PASSWORD", "integrum_pass"),
    dbname   = os.getenv("DB_NAME",     "integrum"),
)

TENANT_ID              = 2    # tenants.id for GIL
GIL_PLANT_ID           = 100  # plants.id for GIL_MH_HYBRID
GIL_PES_ID             = 100  # plant_energy_sources.id (WIND, used as combined total)
GIL_CONSUMPTION_UNIT   = 0    # monthly_banking_settlement.consumption_unit_id (NOT NULL, no FK; 0 = GIL whole-plant)

# Excel TOD slot names → tod_slot_definitions.id (discom_id=2 / Maharashtra)
TOD_SLOT_MAP = {
    "normal":   7,   # NORMAL
    "off-peak": 8,   # OFF_PEAK (Night)
    "peak":     5,   # PEAK (Morning; GIL combines morning+evening into one 'peak')
}

MONTH_ABR = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEPT": 9, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}


def d(v):
    if v is None:
        return Decimal(0)
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal(0)


def detect_month_from_filename(fname: str) -> datetime.date:
    """banking_settlement_APR-25.xlsx → 2025-04-01"""
    m = re.search(r"([A-Z]+)-(\d{2})", os.path.basename(fname).upper())
    if not m:
        raise ValueError(f"Cannot parse month from filename: {fname}")
    mon_str, yr_str = m.group(1), m.group(2)
    month_num = MONTH_ABR.get(mon_str)
    if not month_num:
        raise ValueError(f"Unknown month abbreviation: {mon_str}")
    return datetime.date(2000 + int(yr_str), month_num, 1)


def import_monthly_summary(cur, ws, month_date, dry_run):
    """
    Monthly_Summary sheet → monthly_banking_settlement table.

    Excel columns (0-indexed):
      0: month ('2025-04')
      1: TOD_slot ('normal' / 'off-peak' / 'peak')
      2: Generation_value
      3: allocated_consumption
      4: matched_settlement          (round-1 direct match)
      5: surplus_demand              (unmet demand after round 1)
      6: surplus_generation          (surplus gen after round 1)
      7: surplus_gen_with_banking    (total gen matched by banking rounds)
      8: Slot_Total_Consumption
      9: matched_settlement_daily_tod
     10: surplus_gen_daily_tod
     11: surplus_demand_daily_tod
     12: matched_settlement_intra_monthly
     13: surplus_gen_intra_monthly   (lapsed generation)
     14: surplus_demand_intra_monthly (final unmet → grid import)
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        tod_slot = str(row[1]).strip().lower()
        slot_id  = TOD_SLOT_MAP.get(tod_slot)
        if slot_id is None:
            print(f"  WARNING: unknown TOD slot '{tod_slot}' — skipping")
            continue

        gen_val         = d(row[2])   # Generation_value
        slot_cons       = d(row[8])   # Slot_Total_Consumption
        matched         = d(row[4])   # direct matched (round 1)
        surplus_demand  = d(row[5])   # unmet demand after round 1
        surplus_gen     = d(row[6])   # surplus gen after round 1
        tod_banking     = d(row[9])   # daily TOD banking utilised
        intra_banking   = d(row[12])  # intra-monthly banking utilised
        surp_lapsed     = d(row[13])  # surplus_gen_intra_monthly (lapsed/expired)
        grid_import     = d(row[14])  # surplus_demand_intra_monthly (net unmet → grid)

        banking_utilised = tod_banking + intra_banking
        total_matched    = matched + banking_utilised

        rows.append((
            TENANT_ID,
            GIL_PES_ID,
            GIL_CONSUMPTION_UNIT,
            month_date,
            slot_id,
            gen_val,           # gross_generation_kwh
            Decimal(0),        # generation_losses_kwh (not in Excel)
            gen_val,           # net_generation_kwh
            slot_cons,         # total_consumption_kwh
            matched,           # direct_matched_kwh
            banking_utilised,  # banking_utilised_kwh
            total_matched,     # total_matched_kwh
            Decimal(0),        # opening_banking_balance_kwh
            surplus_gen,       # surplus_before_banking_kwh
            intra_banking,     # intra_month_banking_kwh
            Decimal(0),        # carry_forward_banking_kwh
            Decimal(0),        # banking_loss_kwh
            surp_lapsed,       # surplus_lapsed_kwh
            Decimal(0),        # closing_banking_balance_kwh
            surplus_demand,    # unmet_demand_kwh (after round 1)
            grid_import,       # grid_import_kwh (final unmet after all banking)
            Decimal(0),        # over_injection_kwh
        ))

    if not rows:
        print("  No rows to insert")
        return

    if dry_run:
        print(f"  [DRY] Would insert {len(rows)} monthly_banking_settlement rows")
        return

    # Delete existing rows for this month before re-inserting (avoids functional-index ON CONFLICT)
    cur.execute("""
        DELETE FROM monthly_banking_settlement
        WHERE tenant_id = %s AND plant_energy_source_id = %s
          AND consumption_unit_id = %s AND month = %s
    """, (TENANT_ID, GIL_PES_ID, GIL_CONSUMPTION_UNIT, month_date))

    execute_values(cur, """
        INSERT INTO monthly_banking_settlement (
            tenant_id, plant_energy_source_id, consumption_unit_id, month, tod_slot_id,
            gross_generation_kwh, generation_losses_kwh, net_generation_kwh,
            total_consumption_kwh, direct_matched_kwh, banking_utilised_kwh,
            total_matched_kwh, opening_banking_balance_kwh,
            surplus_before_banking_kwh, intra_month_banking_kwh,
            carry_forward_banking_kwh, banking_loss_kwh, surplus_lapsed_kwh,
            closing_banking_balance_kwh, unmet_demand_kwh,
            grid_import_kwh, over_injection_kwh
        ) VALUES %s
    """, rows)
    print(f"  Inserted {len(rows)} monthly_banking_settlement rows")


def import_tod_wise_summary(cur, ws, dry_run):
    """
    TOD_Wise_Summary sheet → tod_daily_summary table.

    Excel columns (0-indexed):
      0: Date
      1: TOD_slot
      2: Generation_value
      3: allocated_consumption
      4: matched_settlement       (round-1 direct)
      5: surplus_demand
      6: surplus_generation       (before banking)
      7: surplus_gen_with_banking (banking utilised this day)
      8: Slot_Total_Consumption
      9: matched_settlement_daily_tod
     10: surplus_gen_daily_tod
     11: surplus_demand_daily_tod  (→ grid drawl for the day)
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        date_raw = row[0]
        if isinstance(date_raw, str):
            slot_date = datetime.date.fromisoformat(date_raw)
        elif isinstance(date_raw, datetime.datetime):
            slot_date = date_raw.date()
        elif isinstance(date_raw, datetime.date):
            slot_date = date_raw
        else:
            continue

        tod_slot = str(row[1]).strip().lower()
        slot_id  = TOD_SLOT_MAP.get(tod_slot)
        if slot_id is None:
            continue

        gen_val       = d(row[2])   # Generation_value
        slot_cons     = d(row[8])   # Slot_Total_Consumption
        matched       = d(row[4])   # matched_settlement (round 1)
        surplus_gen   = d(row[6])   # surplus_generation (before banking)
        tod_banking   = d(row[9])   # matched_settlement_daily_tod
        grid_drawl    = d(row[11])  # surplus_demand_daily_tod (grid drawl for day)

        total_matched = matched + tod_banking

        rows.append((
            TENANT_ID,
            GIL_PES_ID,
            None,            # consumption_unit_id (nullable in tod_daily_summary)
            slot_date,
            slot_id,
            gen_val,         # generation_kwh
            Decimal(0),      # generation_losses_kwh
            gen_val,         # net_generation_kwh
            slot_cons,       # consumption_kwh
            matched,         # direct_matched_kwh
            tod_banking,     # banking_utilised_kwh
            total_matched,   # total_matched_kwh
            surplus_gen,     # surplus_kwh (before banking)
            Decimal(0),      # lapsed_kwh
            grid_drawl,      # grid_drawl_kwh
        ))

    if not rows:
        print("  No rows to insert")
        return

    if dry_run:
        print(f"  [DRY] Would insert {len(rows)} tod_daily_summary rows")
        return

    # Collect unique dates in this batch, delete before re-inserting
    dates = list({r[3] for r in rows})
    cur.execute("""
        DELETE FROM tod_daily_summary
        WHERE tenant_id = %s AND plant_energy_source_id = %s AND date = ANY(%s)
    """, (TENANT_ID, GIL_PES_ID, dates))

    execute_values(cur, """
        INSERT INTO tod_daily_summary (
            tenant_id, plant_energy_source_id, consumption_unit_id,
            date, tod_slot_id,
            generation_kwh, generation_losses_kwh, net_generation_kwh,
            consumption_kwh, direct_matched_kwh, banking_utilised_kwh,
            total_matched_kwh, surplus_kwh, lapsed_kwh, grid_drawl_kwh
        ) VALUES %s
    """, rows)
    print(f"  Inserted {len(rows)} tod_daily_summary rows")


def import_file(fpath: str, dry_run: bool = False):
    print(f"\n{'='*60}")
    print(f"Importing: {os.path.basename(fpath)}")
    month_date = detect_month_from_filename(fpath)
    print(f"Detected month: {month_date}")
    print(f"{'='*60}")

    if dry_run:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in ["Monthly_Summary", "TOD_Wise_Summary"]:
            if sname in wb.sheetnames:
                rows = list(wb[sname].iter_rows(min_row=2, values_only=True))
                print(f"  [DRY] {sname}: {len(rows)} rows")
        return

    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        print(f"  Using tenant_id={TENANT_ID}, pes_id={GIL_PES_ID}, cu_id={GIL_CONSUMPTION_UNIT}")

        wb = openpyxl.load_workbook(fpath, data_only=True)

        if "Monthly_Summary" in wb.sheetnames:
            print("\n[Monthly_Summary → monthly_banking_settlement]")
            import_monthly_summary(cur, wb["Monthly_Summary"], month_date, dry_run=False)

        if "TOD_Wise_Summary" in wb.sheetnames:
            print("\n[TOD_Wise_Summary → tod_daily_summary]")
            import_tod_wise_summary(cur, wb["TOD_Wise_Summary"], dry_run=False)

        conn.commit()
        print(f"\n✓ Committed all data for {month_date}")

    except Exception as e:
        conn.rollback()
        print(f"\n✗ ERROR — rolled back: {e}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import GIL Excel settlement data")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--dir", default=r"C:\Users\Abcom\AppData\Roaming\Claude\local-agent-mode-sessions\31d015be-27b1-47a2-91d9-339cb003e419\f54f9bc1-6764-456b-a6bb-67d5d37e3317\local_6f45bcb2-5c03-4398-a081-79ff081bddcc\uploads\gil_data\GIL SETTLEMENT_NEW",
                        help="Directory containing banking_settlement_*.xlsx files")
    args = parser.parse_args()

    dirs_to_check = [args.dir, r"D:\Integrum_dashboard\data\gil"]
    files = []
    for d_dir in dirs_to_check:
        if os.path.isdir(d_dir):
            files = sorted(
                os.path.join(d_dir, f) for f in os.listdir(d_dir)
                if f.startswith("banking_settlement_") and f.endswith(".xlsx")
            )
            if files:
                break

    if not files:
        print("No GIL files found. Check --dir argument.")
        sys.exit(1)

    print(f"Found {len(files)} GIL files")
    for fpath in files:
        import_file(fpath, dry_run=args.dry_run)

    print("\nDone.")
