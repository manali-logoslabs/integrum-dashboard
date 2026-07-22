"""
import_c9_excel.py
==================
Imports C9 settlement data from Excel files into the unified schema.

Usage:
    python import_c9_excel.py [--dry-run] [--file path/to/file.xlsx] [--all]

Files expected (place in D:\Integrum_dashboard\data\c9\):
    HRBR_Aug_Gen_Consumption_15min.xlsx     → 2025-08
    Solar_Settlement_September_2025.xlsx    → 2025-09
    Solar_Settlement_November_2025.xlsx     → 2025-11

The script is idempotent: re-running it on the same file will upsert data.
"""

import os, sys, argparse, datetime, re
from decimal import Decimal
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── CONFIG ────────────────────────────────────────────────────────────────────

DB_CONFIG = dict(
    host     = os.getenv("DB_HOST",     "localhost"),
    port     = int(os.getenv("DB_PORT", "5432")),
    user     = os.getenv("DB_USER",     "integrum"),
    password = os.getenv("DB_PASSWORD", "integrum_pass"),
    dbname   = os.getenv("DB_NAME",     "integrum"),
)

TENANT_ID = 1   # C9 tenant

# C9 consumption units master — order matches priority (Group A first, then B)
C9_UNITS = [
    # (unit_code,        unit_name,                                        group, tariff)
    ("MALLESWARAM",     "MALLESWARAM (C2HT-136)",                         "A",   7.20),
    ("OLD_AIRPORT_RD",  "OLD AIRPORT ROAD (E6HT209)",                     "A",   7.20),
    ("SAHAKAR_NAGAR",   "SAHAKAR NAGAR (C8HT-111)",                       "A",   7.20),
    ("HRBR_UNIT",       "HRBR UNIT (E8HT-203)",                           "A",   7.20),
    ("THANISANDRA",     "THANISANDRA (C8HT-135)",                         "B",   5.95),
    ("WHITEFIELD",      "WHITEFIELD (E4HT-355)",                          "B",   5.95),
    ("BELLANDUR_CORP",  "BELLANDUR CORP. OFFICE (S11BHT 406)",            "B",   5.95),
    ("BELLANDUR",       "BELLANDUR (S11HT-124)",                          "B",   5.95),
    ("SARJAPURA",       "SARJAPURA (S11HT-419)",                          "B",   5.95),
    ("KANAKAPURA",      "KANAKAPURA (S12HT-99)",                          "B",   5.95),
    ("ELECTRONIC_CITY", "ELECTRONIC CITY (S13HT-87)",                     "B",   5.95),
    ("SLOT_SURPLUS",    "Slot_Surplus",                                   "X",   0.00),  # virtual
]

# Excel unit_name → unit_code mapping (handles minor name variations)
UNIT_NAME_MAP = {
    "MALLESWARAM (C2HT-136)":                    "MALLESWARAM",
    "OLD AIRPORT ROAD (E6HT209)":                "OLD_AIRPORT_RD",
    "SAHAKAR NAGAR (C8HT-111)":                  "SAHAKAR_NAGAR",
    "HRBR UNIT (E8HT-203)":                      "HRBR_UNIT",
    "THANISANDRA (C8HT-135)":                    "THANISANDRA",
    "WHITEFIELD (E4HT-355)":                     "WHITEFIELD",
    "BELLANDUR CORP. OFFICE (S11BHT 406)":       "BELLANDUR_CORP",
    "BELLANDUR (S11HT-124)":                     "BELLANDUR",
    "SARJAPURA (S11HT-419)":                     "SARJAPURA",
    "KANAKAPURA (S12HT-99)":                     "KANAKAPURA",
    "ELECTRONIC CITY (S13HT-87)":               "ELECTRONIC_CITY",
    "Slot_Surplus":                               "SLOT_SURPLUS",
    "GRAND TOTAL":                               None,  # skip
}

# Gen_Consumption_15min column headers → unit_code
GEN_CONS_UNIT_COLS = {
    "MALLESWARAM (C2HT-136)":             "MALLESWARAM",
    "OLD AIRPORT ROAD (E6HT209)":         "OLD_AIRPORT_RD",
    "SAHAKAR NAGAR (C8HT-111)":           "SAHAKAR_NAGAR",
    "HRBR UNIT (E8HT-203)":              "HRBR_UNIT",
    "THANISANDRA (C8HT-135)":            "THANISANDRA",
    "WHITEFIELD (E4HT-355)":             "WHITEFIELD",
    "BELLANDUR CORP. OFFICE (S11BHT 406)": "BELLANDUR_CORP",
    "BELLANDUR (S11HT-124)":             "BELLANDUR",
    "SARJAPURA (S11HT-419)":             "SARJAPURA",
    "KANAKAPURA (S12HT-99)":             "KANAKAPURA",
    "ELECTRONIC CITY (S13HT-87)":        "ELECTRONIC_CITY",
}


# ── HELPERS ───────────────────────────────────────────────────────────────────

def d(v):
    """Coerce to Decimal, default 0."""
    if v is None:
        return Decimal(0)
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal(0)


def ensure_units(cur, dry_run=False):
    """Upsert consumption_units for C9 tenant. Returns {unit_code: unit_id}."""
    mapping = {}
    for code, name, group, rate in C9_UNITS:
        if not dry_run:
            cur.execute("""
                INSERT INTO consumption_units (tenant_id, unit_code, unit_name, tariff_group, tariff_rate)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (tenant_id, unit_code) DO UPDATE
                    SET unit_name    = EXCLUDED.unit_name,
                        tariff_group = EXCLUDED.tariff_group,
                        tariff_rate  = EXCLUDED.tariff_rate
                RETURNING unit_id
            """, (TENANT_ID, code, name, group, rate))
            uid = cur.fetchone()[0]
        else:
            uid = None
        mapping[code] = uid
    return mapping


def unit_id_from_name(unit_map_code, code_to_id):
    """Resolve unit_name → unit_id. Returns None to skip."""
    code = UNIT_NAME_MAP.get(unit_map_code)
    if code is None:
        return None
    return code_to_id.get(code)


# ── SHEET IMPORTERS ──────────────────────────────────────────────────────────

def import_summary(cur, ws, month_date, dry_run):
    """Summary sheet → c9_monthly_summary."""
    data = {}
    key_map = {
        "Total Generation (kWh)":                "total_generation_kwh",
        "Allocated to Units (kWh)":              "allocated_to_units_kwh",
        "Raw Slot Surplus — gross (kWh)":        "raw_slot_surplus_kwh",
        "Banking Charge Lost (8%) (kWh)":        "banking_charge_lost_kwh",
        "Net Banked — after charge (kWh)":       "net_banked_kwh",
        "Total Consumption (kWh)":               "total_consumption_kwh",
        "Round 1 Matched Settlement (kWh)":      "round1_matched_kwh",
        "Round 1 Surplus Demand (kWh)":          "round1_surplus_demand_kwh",
        "Round 2 Matched from Bank (kWh)":       "round2_matched_kwh",
        "Lapse Units (kWh)":                     "lapse_units_kwh",
        "Final Grid Consumption (kWh)":          "final_grid_consumption_kwh",
    }
    for row in ws.iter_rows(values_only=True):
        if row[0] in key_map and row[1] is not None:
            data[key_map[row[0]]] = d(row[1])

    # Aug has no structured Summary with these keys — derive from Unit_Wise_Monthly instead
    # For Aug, we may only have total_generation from row "Total Generation (kWh)"
    if not dry_run and data:
        cur.execute("""
            INSERT INTO c9_monthly_summary (
                tenant_id, month,
                total_generation_kwh, allocated_to_units_kwh,
                raw_slot_surplus_kwh, banking_charge_lost_kwh,
                net_banked_kwh, total_consumption_kwh,
                round1_matched_kwh, round1_surplus_demand_kwh,
                round2_matched_kwh, lapse_units_kwh,
                final_grid_consumption_kwh
            ) VALUES (
                %(tid)s, %(m)s,
                %(tgen)s, %(alloc)s,
                %(raw)s, %(bloss)s,
                %(net)s, %(tcons)s,
                %(r1m)s, %(r1sd)s,
                %(r2m)s, %(lapse)s,
                %(grid)s
            )
            ON CONFLICT (tenant_id, month) DO UPDATE SET
                total_generation_kwh       = EXCLUDED.total_generation_kwh,
                allocated_to_units_kwh     = EXCLUDED.allocated_to_units_kwh,
                raw_slot_surplus_kwh       = EXCLUDED.raw_slot_surplus_kwh,
                banking_charge_lost_kwh    = EXCLUDED.banking_charge_lost_kwh,
                net_banked_kwh             = EXCLUDED.net_banked_kwh,
                total_consumption_kwh      = EXCLUDED.total_consumption_kwh,
                round1_matched_kwh         = EXCLUDED.round1_matched_kwh,
                round1_surplus_demand_kwh  = EXCLUDED.round1_surplus_demand_kwh,
                round2_matched_kwh         = EXCLUDED.round2_matched_kwh,
                lapse_units_kwh            = EXCLUDED.lapse_units_kwh,
                final_grid_consumption_kwh = EXCLUDED.final_grid_consumption_kwh
        """, dict(
            tid=TENANT_ID, m=month_date,
            tgen=data.get("total_generation_kwh", 0),
            alloc=data.get("allocated_to_units_kwh", 0),
            raw=data.get("raw_slot_surplus_kwh", 0),
            bloss=data.get("banking_charge_lost_kwh", 0),
            net=data.get("net_banked_kwh", 0),
            tcons=data.get("total_consumption_kwh", 0),
            r1m=data.get("round1_matched_kwh", 0),
            r1sd=data.get("round1_surplus_demand_kwh", 0),
            r2m=data.get("round2_matched_kwh", 0),
            lapse=data.get("lapse_units_kwh", 0),
            grid=data.get("final_grid_consumption_kwh", 0),
        ))


def import_gen_cons_15min(cur, ws, month_date, dry_run):
    """Gen_Consumption_15min sheet → c9_slot_generation + c9_slot_consumption."""
    headers = [cell for cell in next(ws.iter_rows(values_only=True))]
    # headers[0] = DateTime, headers[1] = Generation_kWh, headers[2..] = units
    gen_rows = []
    cons_rows_by_col = {i: [] for i in range(2, len(headers)) if headers[i] in GEN_CONS_UNIT_COLS}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        dt = row[0]
        if isinstance(dt, str):
            dt = datetime.datetime.fromisoformat(dt)
        # IST → UTC offset; store as naive UTC (or with tz)
        slot_ts = dt.replace(tzinfo=datetime.timezone(datetime.timedelta(hours=5, minutes=30)))
        gen_kwh = d(row[1])
        gen_rows.append((TENANT_ID, slot_ts, gen_kwh))
        for col_idx in cons_rows_by_col:
            unit_col_name = headers[col_idx]
            unit_code = GEN_CONS_UNIT_COLS.get(unit_col_name)
            if unit_code:
                cons_rows_by_col[col_idx].append((unit_code, slot_ts, d(row[col_idx])))

    if not dry_run:
        # generation
        execute_values(cur, """
            INSERT INTO c9_slot_generation (tenant_id, slot_ts, generation_kwh)
            VALUES %s
            ON CONFLICT (tenant_id, slot_ts) DO UPDATE SET generation_kwh = EXCLUDED.generation_kwh
        """, gen_rows, template="(%s, %s, %s)")
        print(f"  Inserted {len(gen_rows)} generation slots")

        # consumption — need unit_ids
        for col_idx, rows in cons_rows_by_col.items():
            unit_col_name = headers[col_idx]
            unit_code = GEN_CONS_UNIT_COLS.get(unit_col_name)
            unit_id = _unit_id_cache.get(unit_code)
            if not unit_id:
                continue
            data = [(TENANT_ID, unit_id, slot_ts, kwh) for (_, slot_ts, kwh) in rows]
            execute_values(cur, """
                INSERT INTO c9_slot_consumption (tenant_id, unit_id, slot_ts, consumption_kwh)
                VALUES %s
                ON CONFLICT (tenant_id, unit_id, slot_ts) DO UPDATE SET consumption_kwh = EXCLUDED.consumption_kwh
            """, data, template="(%s, %s, %s, %s)")
        print(f"  Inserted consumption for {len(cons_rows_by_col)} units")
    else:
        print(f"  [DRY] Would insert {len(gen_rows)} generation slots + consumption")


def import_tod_unit_wise_monthly(cur, ws, month_date, dry_run):
    """TOD_Unit_Wise_Monthly sheet → c9_monthly_tod."""
    rows_to_insert = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[0] == "Unit_Name":
            continue
        unit_name = str(row[0]).strip()
        unit_code = UNIT_NAME_MAP.get(unit_name)
        if unit_code is None or unit_code == "SLOT_SURPLUS":
            continue
        unit_id = _unit_id_cache.get(unit_code)
        if not unit_id:
            print(f"  WARNING: unknown unit '{unit_name}'")
            continue
        tod_slot = str(row[1]).strip() if row[1] else "Unknown"
        rows_to_insert.append((
            TENANT_ID, unit_id, month_date, tod_slot,
            d(row[2]), d(row[3]), d(row[4]), d(row[5]), d(row[6])
        ))
    if not dry_run and rows_to_insert:
        execute_values(cur, """
            INSERT INTO c9_monthly_tod
                (tenant_id, unit_id, month, tod_slot,
                 allocated_generation, consumption_kwh, matched_settlement,
                 surplus_generation, surplus_demand)
            VALUES %s
            ON CONFLICT (tenant_id, unit_id, month, tod_slot) DO UPDATE SET
                allocated_generation = EXCLUDED.allocated_generation,
                consumption_kwh      = EXCLUDED.consumption_kwh,
                matched_settlement   = EXCLUDED.matched_settlement,
                surplus_generation   = EXCLUDED.surplus_generation,
                surplus_demand       = EXCLUDED.surplus_demand
        """, rows_to_insert, template="(%s,%s,%s,%s,%s,%s,%s,%s,%s)")
        print(f"  Inserted {len(rows_to_insert)} monthly TOD rows")
    else:
        print(f"  [DRY] Would insert {len(rows_to_insert)} monthly TOD rows")


def import_unit_wise_monthly(cur, ws, month_date, dry_run):
    """Unit_Wise_Monthly sheet → c9_unit_monthly.
    Handles both Aug format (10 cols) and Sep/Nov format (9 cols).
    """
    headers = list(next(ws.iter_rows(values_only=True)))
    has_gen_without_banking = "Generation_Without_Banking" in headers

    rows_to_insert = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        unit_name = str(row[0]).strip()
        unit_code = UNIT_NAME_MAP.get(unit_name)
        if unit_code is None:  # GRAND TOTAL — skip
            continue
        unit_id = _unit_id_cache.get(unit_code)
        if not unit_id:
            print(f"  WARNING: unknown unit '{unit_name}'")
            continue

        if has_gen_without_banking:
            # Aug: cols = Unit_Name, Alloc_Gen, Consumption, Matched, Surplus_Gen,
            #            Surplus_Demand, Gen_Without_Banking, Matched_2, Lapse, Grid
            alloc_gen  = d(row[1])
            cons       = d(row[2])
            matched    = d(row[3])
            surplus_g  = d(row[4])
            surplus_d  = d(row[5])
            # row[6] = Generation_Without_Banking — informational, skip
            matched_2  = d(row[7])
            lapse      = d(row[8])
            grid_cons  = d(row[9])
        else:
            # Sep/Nov: cols = Unit_Name, Alloc_Gen, Consumption, Matched, Surplus_Gen,
            #                 Surplus_Demand, Matched_2, Lapse, Grid
            alloc_gen  = d(row[1])
            cons       = d(row[2])
            matched    = d(row[3])
            surplus_g  = d(row[4])
            surplus_d  = d(row[5])
            matched_2  = d(row[6])
            lapse      = d(row[7])
            grid_cons  = d(row[8])

        rows_to_insert.append((
            TENANT_ID, unit_id, month_date,
            alloc_gen, cons, matched, surplus_g, surplus_d,
            matched_2, lapse, grid_cons
        ))

    if not dry_run and rows_to_insert:
        execute_values(cur, """
            INSERT INTO c9_unit_monthly
                (tenant_id, unit_id, month,
                 allocated_generation, consumption_kwh, matched_settlement,
                 surplus_generation, surplus_demand,
                 matched_settlement_2, lapse_units, grid_consumption)
            VALUES %s
            ON CONFLICT (tenant_id, unit_id, month) DO UPDATE SET
                allocated_generation = EXCLUDED.allocated_generation,
                consumption_kwh      = EXCLUDED.consumption_kwh,
                matched_settlement   = EXCLUDED.matched_settlement,
                surplus_generation   = EXCLUDED.surplus_generation,
                surplus_demand       = EXCLUDED.surplus_demand,
                matched_settlement_2 = EXCLUDED.matched_settlement_2,
                lapse_units          = EXCLUDED.lapse_units,
                grid_consumption     = EXCLUDED.grid_consumption
        """, rows_to_insert, template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)")
        print(f"  Inserted {len(rows_to_insert)} unit monthly rows")
    else:
        print(f"  [DRY] Would insert {len(rows_to_insert)} unit monthly rows")


# ── MONTH DETECTION ──────────────────────────────────────────────────────────

def detect_month(wb) -> datetime.date:
    """Auto-detect month from Summary or file content."""
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for row in ws.iter_rows(values_only=True):
            if row[0] and ("Settlement Summary" in str(row[0]) or "Generation vs" in str(row[0])):
                text = str(row[0])
                # e.g. "Solar Settlement Summary — 2025-09"  or "... August 2025"
                m = re.search(r"(\d{4})-(\d{2})", text)
                if m:
                    return datetime.date(int(m.group(1)), int(m.group(2)), 1)
                m2 = re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})", text, re.IGNORECASE)
                if m2:
                    month_num = ["january","february","march","april","may","june","july",
                                 "august","september","october","november","december"].index(m2.group(1).lower()) + 1
                    return datetime.date(int(m2.group(2)), month_num, 1)
    raise ValueError("Could not detect month from file")


# ── MAIN IMPORT ──────────────────────────────────────────────────────────────

_unit_id_cache: dict = {}


def import_file(fpath: str, dry_run: bool = False):
    print(f"\n{'='*60}")
    print(f"Importing: {os.path.basename(fpath)}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(fpath, data_only=True)
    month_date = detect_month(wb)
    print(f"Detected month: {month_date}")

    if dry_run:
        print("[DRY RUN — no DB writes]")
        # Still parse to count rows
        if "Unit_Wise_Monthly" in wb.sheetnames:
            import_unit_wise_monthly(None, wb["Unit_Wise_Monthly"], month_date, dry_run=True)
        if "TOD_Unit_Wise_Monthly" in wb.sheetnames:
            import_tod_unit_wise_monthly(None, wb["TOD_Unit_Wise_Monthly"], month_date, dry_run=True)
        return

    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    try:
        cur = conn.cursor()

        # Ensure units exist & build cache
        global _unit_id_cache
        _unit_id_cache = ensure_units(cur, dry_run=False)

        # 1. Summary KPIs
        if "Summary" in wb.sheetnames:
            print("\n[Summary]")
            import_summary(cur, wb["Summary"], month_date, dry_run=False)

        # 2. 15-min generation + consumption
        if "Gen_Consumption_15min" in wb.sheetnames:
            print("\n[Gen_Consumption_15min]")
            import_gen_cons_15min(cur, wb["Gen_Consumption_15min"], month_date, dry_run=False)

        # 3. Monthly TOD breakdown
        if "TOD_Unit_Wise_Monthly" in wb.sheetnames:
            print("\n[TOD_Unit_Wise_Monthly]")
            import_tod_unit_wise_monthly(cur, wb["TOD_Unit_Wise_Monthly"], month_date, dry_run=False)

        # 4. Monthly unit totals (the most critical table)
        if "Unit_Wise_Monthly" in wb.sheetnames:
            print("\n[Unit_Wise_Monthly]")
            import_unit_wise_monthly(cur, wb["Unit_Wise_Monthly"], month_date, dry_run=False)

        conn.commit()
        print(f"\n✓ Committed all data for {month_date}")

    except Exception as e:
        conn.rollback()
        print(f"\n✗ ERROR — rolled back: {e}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import C9 Excel settlement data")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")
    parser.add_argument("--file", type=str, help="Import a single file")
    parser.add_argument("--all", action="store_true", help="Import all files in data/c9/")
    parser.add_argument("--data-dir", default=r"D:\Integrum_dashboard\data\c9",
                        help="Directory containing C9 Excel files")
    args = parser.parse_args()

    files_to_import = []

    if args.file:
        files_to_import = [args.file]
    elif args.all:
        d_dir = args.data_dir
        files_to_import = [
            os.path.join(d_dir, f) for f in os.listdir(d_dir)
            if f.endswith(".xlsx")
        ]
    else:
        # Default: import the 3 known files from uploads
        default_files = [
            r"C:\Users\Abcom\AppData\Roaming\Claude\local-agent-mode-sessions\31d015be-27b1-47a2-91d9-339cb003e419\f54f9bc1-6764-456b-a6bb-67d5d37e3317\local_6f45bcb2-5c03-4398-a081-79ff081bddcc\uploads\HRBR_Aug_Gen_Consumption_15min.xlsx",
            r"C:\Users\Abcom\AppData\Roaming\Claude\local-agent-mode-sessions\31d015be-27b1-47a2-91d9-339cb003e419\f54f9bc1-6764-456b-a6bb-67d5d37e3317\local_6f45bcb2-5c03-4398-a081-79ff081bddcc\uploads\Solar_Settlement_September_2025.xlsx",
            r"C:\Users\Abcom\AppData\Roaming\Claude\local-agent-mode-sessions\31d015be-27b1-47a2-91d9-339cb003e419\f54f9bc1-6764-456b-a6bb-67d5d37e3317\local_6f45bcb2-5c03-4398-a081-79ff081bddcc\uploads\Solar_Settlement_November_2025.xlsx",
        ]
        files_to_import = [f for f in default_files if os.path.exists(f)]

    if not files_to_import:
        print("No files found. Use --file <path> or --all --data-dir <dir>")
        sys.exit(1)

    for fpath in sorted(files_to_import):
        import_file(fpath, dry_run=args.dry_run)

    print("\nDone.")
